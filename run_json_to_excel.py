#!/usr/bin/env python3
"""Export business_extract JSON files to an Excel workbook.

Each row is keyed by one sub_project_id. File-level fields are repeated for
all sub-project rows in the same JSON.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any


def json_cell(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def normalize_unit(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\$\s*m\^\{?2\}?\s*\$", "m2", text)
    text = re.sub(r"\s+", "", text)
    return text


def is_detail_sub_item_row(row: dict[str, Any]) -> bool:
    seq = str(row.get("seq") or "").strip()
    project_name = str(row.get("project_name") or "").strip()
    if not seq or not re.fullmatch(r"\d+(?:\.\d+)?", seq):
        return False
    return bool(project_name)


def quality_warnings(row: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    project_code = str(row.get("project_code") or "").strip()
    project_name = str(row.get("project_name") or "").strip()
    unit_price = str(row.get("unit_price") or "").strip()
    sub_project_id = str(row.get("sub_project_id") or "").strip()
    combined = f"{project_code} {project_name} {sub_project_id}"

    if not project_code:
        warnings.append("缺少code")
    if unit_price and (unit_price == project_code or re.fullmatch(r"\d{9,15}", unit_price)):
        warnings.append("疑似金额列错位")
    if any(token in project_name for token in ("本页小计", "合计", "小计")):
        warnings.append("疑似小计/合计")
    if re.search(r"-X\d+(?:-\d+){2,}", combined):
        warnings.append("疑似坐标乱码")
    return warnings


def cleaned_sub_item_rows(rows: Any) -> list[dict[str, Any]]:
    if not isinstance(rows, list):
        return []

    cleaned_rows: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict) or not is_detail_sub_item_row(row):
            continue
        cleaned_row = dict(row)
        cleaned_row["unit"] = normalize_unit(cleaned_row.get("unit"))
        cleaned_rows.append(cleaned_row)
    return cleaned_rows


def sub_item_rows_cell(rows: Any) -> str:
    return json_cell(cleaned_sub_item_rows(rows))


def sub_item_warnings_cell(rows: Any) -> str:
    warnings: list[dict[str, Any]] = []
    for index, row in enumerate(cleaned_sub_item_rows(rows)):
        row_warnings = quality_warnings(row)
        if not row_warnings:
            continue
        warnings.append({
            "row_index": index,
            "seq": row.get("seq", ""),
            "project_code": row.get("project_code", ""),
            "project_name": row.get("project_name", ""),
            "warnings": row_warnings,
        })
    return json_cell(warnings)


def construction_process_content(processes: Any) -> str:
    if not isinstance(processes, list):
        return ""
    contents = []
    for item in processes:
        if isinstance(item, dict):
            content = str(item.get("content") or "").strip()
            if content:
                contents.append(content)
    return "\n\n---\n\n".join(contents)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        raise ValueError(f"JSON root is not an object: {path}")
    return data


def discover_json_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    preferred = sorted(input_path.rglob("business_extract.json"))
    if preferred:
        return preferred
    return sorted(input_path.rglob("*.json"))


def rows_from_payload(payload: dict[str, Any], source_json: Path) -> list[dict[str, Any]]:
    document_info = payload.get("document_info")
    if not isinstance(document_info, dict):
        document_info = {}

    shared = {
        "file_name": payload.get("file_name", ""),
        "consultation_project_name": document_info.get("consultation_project_name", ""),
        "consultation_time": document_info.get("consultation_time", "") or document_info.get("consultation_date", ""),
        "location": document_info.get("location", "") or payload.get("location", ""),
        "renovation_content": document_info.get("renovation_content", ""),
    }

    sub_projects = payload.get("sub_projects")
    if not isinstance(sub_projects, list) or not sub_projects:
        return [
            {
                **shared,
                "sub_item_project_rows": "",
                "sub_item_project_warnings": "",
            }
        ]

    rows: list[dict[str, Any]] = []
    for sub_project in sub_projects:
        if not isinstance(sub_project, dict):
            continue
        rows.append(
            {
                **shared,
                "sub_item_project_rows": sub_item_rows_cell(sub_project.get("sub_item_project_rows", [])),
                "sub_item_project_warnings": sub_item_warnings_cell(sub_project.get("sub_item_project_rows", [])),
            }
        )
    return rows


def write_excel(rows: list[dict[str, Any]], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install it with `pip install openpyxl`."
        ) from exc

    headers = [
        "file_name",
        "consultation_project_name",
        "consultation_time",
        "location",
        "renovation_content",
        "sub_item_project_rows",
        "sub_item_project_warnings",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "by_sub_project"
    sheet.append(headers)

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrap_columns = {
        "consultation_project_name",
        "renovation_content",
        "sub_item_project_rows",
        "sub_item_project_warnings",
    }
    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        if header in wrap_columns:
            sheet.column_dimensions[letter].width = 42
        else:
            sheet.column_dimensions[letter].width = 28
        for cell in sheet[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert business_extract JSON files to Excel.")
    parser.add_argument("--input", required=True, help="A business_extract.json file or a directory.")
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    json_files = discover_json_files(input_path)
    if not json_files:
        print(f"No JSON files found in {input_path}")
        sys.exit(1)

    all_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    for json_file in json_files:
        try:
            payload = load_json(json_file)
            all_rows.extend(rows_from_payload(payload, json_file))
        except Exception as exc:
            errors.append(f"{json_file}: {exc}")

    if not all_rows:
        print("No rows exported.")
        for error in errors:
            print(f"ERROR: {error}")
        sys.exit(1)

    write_excel(all_rows, output_path)
    print(f"Exported {len(all_rows)} rows from {len(json_files)} JSON file(s): {output_path}")
    if errors:
        print(f"Skipped {len(errors)} file(s) with errors:")
        for error in errors:
            print(f"  - {error}")


if __name__ == "__main__":
    main()
