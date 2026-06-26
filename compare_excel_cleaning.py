#!/usr/bin/env python3
"""Compare raw sub-item rows with the exported Excel rows."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import csv
import json
import re
from pathlib import Path
from typing import Any


CORE_ROW_FIELDS = [
    "seq",
    "project_code",
    "project_name",
    "project_description",
    "unit",
    "quantity",
    "unit_price",
    "total_price",
    "provisional_estimate",
    "labor_cost",
    "machinery_cost",
    "remark",
]


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        raise ValueError(f"JSON root is not object: {path}")
    return data


def resolve_business_root(input_path: Path) -> Path:
    input_path = input_path.expanduser().resolve()
    if input_path.name == "business_json_vl_llm":
        return input_path
    nested = input_path / "business_json_vl_llm"
    if nested.exists():
        return nested
    nested = input_path / "audit_ocr_vl_llm_json" / "business_json_vl_llm"
    if nested.exists():
        return nested
    return input_path


def is_blank(value: Any) -> bool:
    return str(value or "").strip() == ""


def is_raw_empty_row(row: dict[str, Any]) -> bool:
    return all(is_blank(row.get(field)) for field in CORE_ROW_FIELDS)


def is_valid_seq(value: Any) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d+)?", str(value or "").strip()))


def normalize_unit(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\$\s*m\^\{?2\}?\s*\$", "m2", text)
    text = re.sub(r"\s+", "", text)
    return text


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    normalized["unit"] = normalize_unit(normalized.get("unit"))
    return normalized


def row_key(row: dict[str, Any]) -> str:
    return json.dumps(normalize_row(row), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def removal_reasons(row: dict[str, Any]) -> list[str]:
    reasons: list[str] = []
    if not is_valid_seq(row.get("seq")):
        reasons.append("缺少seq")
    if is_blank(row.get("project_name")):
        reasons.append("缺少name")
    return reasons or ["其他"]


def quality_warnings(row: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    project_code = str(row.get("project_code") or "").strip()
    project_name = str(row.get("project_name") or "").strip()
    unit_price = str(row.get("unit_price") or "").strip()
    sub_project_id = str(row.get("sub_project_id") or "").strip()
    combined = f"{project_code} {project_name} {sub_project_id}"

    if not project_code:
        warnings.append("缺少code")
    if unit_price and (unit_price == project_code or re.fullmatch(r"\d{9,15}", unit_price)):
        warnings.append("疑似金额列错位")
    if any(token in project_name for token in ("本页小计", "合计", "小计")):
        warnings.append("疑似小计/合计")
    if re.search(r"-X\d+(?:-\d+){2,}", combined):
        warnings.append("疑似坐标乱码")
    return warnings


def row_preview(row: dict[str, Any]) -> str:
    fields = [
        row.get("seq"),
        row.get("project_code"),
        row.get("project_name"),
        row.get("project_description"),
        row.get("unit"),
        row.get("quantity"),
        row.get("unit_price"),
        row.get("total_price"),
    ]
    return " | ".join(str(value or "").strip() for value in fields).strip(" |")


def iter_sub_item_rows(payload: dict[str, Any]) -> list[tuple[int, str, int, dict[str, Any]]]:
    result: list[tuple[int, str, int, dict[str, Any]]] = []
    sub_projects = payload.get("sub_projects")
    if not isinstance(sub_projects, list):
        return result
    for sp_index, sub_project in enumerate(sub_projects):
        if not isinstance(sub_project, dict):
            continue
        sub_project_id = str(sub_project.get("sub_project_id") or sub_project.get("sub_project_name") or "")
        rows = sub_project.get("sub_item_project_rows")
        if not isinstance(rows, list):
            continue
        for row_index, row in enumerate(rows):
            if isinstance(row, dict):
                result.append((sp_index, sub_project_id, row_index, row))
    return result


def load_excel_rows(excel_path: Path) -> dict[str, Counter[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install it with `pip install openpyxl`.") from exc

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}
    headers = [str(value or "").strip() for value in header_row]
    try:
        file_name_index = headers.index("file_name")
        rows_index = headers.index("sub_item_project_rows")
    except ValueError as exc:
        raise ValueError("Excel must contain file_name and sub_item_project_rows columns.") from exc

    result: dict[str, Counter[str]] = defaultdict(Counter)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        file_name = str(row[file_name_index] or "").strip()
        cell_value = row[rows_index]
        if not file_name or not cell_value:
            continue
        try:
            sub_rows = json.loads(str(cell_value))
        except json.JSONDecodeError:
            continue
        if not isinstance(sub_rows, list):
            continue
        for sub_row in sub_rows:
            if isinstance(sub_row, dict):
                result[file_name][row_key(sub_row)] += 1
    return result


def should_keep_without_excel(row: dict[str, Any]) -> bool:
    return is_valid_seq(row.get("seq")) and not is_blank(row.get("project_name"))


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare raw JSON rows with exported Excel rows.")
    parser.add_argument("--input", dest="input", help="business_json_vl_llm dir or audit_ocr_vl_llm_json dir.")
    parser.add_argument("--json-input", dest="json_input", help="Same as --input. Kept for clearer Excel-vs-JSON commands.")
    parser.add_argument("--excel", help="Current exported Excel file. When provided, compare actual Excel content with raw JSON.")
    parser.add_argument("--summary-output", default="cleaning_compare_summary.tsv", help="Output summary TSV.")
    parser.add_argument("--removed-output", default="cleaning_removed_rows.json", help="Removed row details JSON.")
    parser.add_argument("--removed-tsv-output", default="cleaning_removed_rows.tsv", help="Removed row details TSV.")
    args = parser.parse_args()

    input_arg = args.json_input or args.input
    if not input_arg:
        parser.error("one of --input or --json-input is required")

    business_root = resolve_business_root(Path(input_arg))
    json_files = sorted(business_root.rglob("business_extract.json"))
    excel_rows = load_excel_rows(Path(args.excel).expanduser().resolve()) if args.excel else None
    summaries: list[dict[str, Any]] = []
    removed_details: list[dict[str, Any]] = []

    for json_file in json_files:
        try:
            payload = load_json(json_file)
        except Exception as exc:
            summaries.append({
                "文件名": json_file.parent.name,
                "原始行数": 0,
                "Excel行数": 0,
                "被删除行数": 0,
                "缺少序号行数": 0,
                "缺少name行数": 0,
                "保留但缺少code行数": 0,
                "备注": f"JSON读取失败:{exc}",
                "JSON路径": str(json_file),
            })
            continue

        rows = iter_sub_item_rows(payload)
        file_name = str(payload.get("file_name") or json_file.parent.name)
        remaining_excel = Counter(excel_rows.get(file_name, Counter())) if excel_rows is not None else Counter()
        excel_count = sum(remaining_excel.values()) if excel_rows is not None else 0
        kept = 0
        missing_seq = 0
        missing_name = 0
        kept_missing_code = 0

        for sp_index, sub_project_id, row_index, row in rows:
            if excel_rows is None:
                is_removed = not should_keep_without_excel(row)
            else:
                key = row_key(row)
                is_removed = remaining_excel[key] <= 0
                if not is_removed:
                    remaining_excel[key] -= 1

            if not is_removed:
                kept += 1
                if is_blank(row.get("project_code")):
                    kept_missing_code += 1
                continue

            reasons = removal_reasons(row)
            if "缺少seq" in reasons:
                missing_seq += 1
            if "缺少name" in reasons:
                missing_name += 1
            warnings = quality_warnings(row)

            removed_details.append({
                "文件名": file_name,
                "子项目序号": sp_index,
                "子项目": sub_project_id,
                "行序号": row_index,
                "页码": row.get("page_no"),
                "删除原因": "、".join(reasons),
                "质量提示": "、".join(warnings),
                "缺少seq": "是" if "缺少seq" in reasons else "否",
                "缺少name": "是" if "缺少name" in reasons else "否",
                "缺少code": "是" if "缺少code" in warnings else "否",
                "预览": row_preview(row),
                "原始行": row,
                "JSON路径": str(json_file),
            })

        removed = len(rows) - kept
        summaries.append({
            "文件名": file_name,
            "原始行数": len(rows),
            "Excel行数": excel_count if excel_rows is not None else kept,
            "被删除行数": removed,
            "缺少序号行数": missing_seq,
            "缺少name行数": missing_name,
            "保留但缺少code行数": kept_missing_code,
            "备注": "",
            "JSON路径": str(json_file),
        })

    summary_path = Path(args.summary_output).expanduser()
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "文件名",
        "原始行数",
        "Excel行数",
        "被删除行数",
        "缺少序号行数",
        "缺少name行数",
        "保留但缺少code行数",
        "备注",
        "JSON路径",
    ]
    with summary_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers, delimiter="\t")
        writer.writeheader()
        writer.writerows(summaries)

    removed_path = Path(args.removed_output).expanduser()
    removed_path.parent.mkdir(parents=True, exist_ok=True)
    removed_path.write_text(json.dumps(removed_details, ensure_ascii=False, indent=2), encoding="utf-8")

    removed_tsv_path = Path(args.removed_tsv_output).expanduser()
    removed_tsv_path.parent.mkdir(parents=True, exist_ok=True)
    removed_headers = [
        "文件名",
        "子项目序号",
        "子项目",
        "行序号",
        "页码",
        "删除原因",
        "质量提示",
        "缺少seq",
        "缺少name",
        "缺少code",
        "预览",
        "JSON路径",
    ]
    with removed_tsv_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=removed_headers, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(removed_details)

    print(f"业务JSON数量: {len(json_files)}")
    print(f"对比汇总: {summary_path}")
    print(f"被删除行明细JSON: {removed_path}")
    print(f"被删除行明细TSV: {removed_tsv_path}")
    print(f"被删除行数: {len(removed_details)}")


if __name__ == "__main__":
    main()
